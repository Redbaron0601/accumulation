from scipy.optimize import fsolve
import numpy as np
import math
import pandas as pd
import openpyxl as op

# ['503', '294', '237', '233', '598', '561', '485', '230', '204']
# p_A = (11392.9607416196, 56973.0182393612, 4097.85801775604)
# p_B = (21191.7729307933, 57198.4754529258, 6851.28419162468)
# p_C = (32310.2959656039, 58618.7747140958, 5213.96372217605)
# C is considered whether to choose

# eAB = (p_A[0]-p_B[0], p_A[1]-p_B[1], p_A[2]-p_B[2])
# eAC = (p_A[0]-p_C[0], p_A[1]-p_C[1], p_A[2]-p_C[2])
#
# m_AB = math.sqrt(math.pow(eAB[0], 2) + \
#                  math.pow(eAB[1], 2) + \
#                  math.pow(eAB[2], 2))
#
# m_AC = math.sqrt(math.pow(eAC[0], 2) + \
#                  math.pow(eAC[1], 2) + \
#                  math.pow(eAC[2], 2))
#
# eAB_norm = (eAB[0]/m_AB, eAB[1]/m_AB, eAB[2]/m_AB)
# eAC_norm = (eAC[0]/m_AC, eAC[1]/m_AC, eAC[2]/m_AC)
# e_vertical = np.cross(eAB_norm, eAC_norm)
# eBO = np.cross(e_vertical, eAB_norm)
# p_O = (p_B[0]+200*eBO[0], p_B[1]+200*eBO[1], p_B[2]+200*eBO[2])


def calc_O_coord(p_A, p_B, p_C):
    eAB = (p_B[0]-p_A[0], p_B[1]-p_A[1], p_B[2]-p_A[2])
    eAC = (p_C[0]-p_A[0], p_C[1]-p_A[1], p_C[2]-p_A[2])
    eBC = (p_C[0]-p_B[0], p_C[1]-p_B[1], p_C[2]-p_B[2])

    m_AB = math.sqrt(math.pow(eAB[0], 2) + \
                     math.pow(eAB[1], 2) + \
                     math.pow(eAB[2], 2))

    m_AC = math.sqrt(math.pow(eAC[0], 2) + \
                     math.pow(eAC[1], 2) + \
                     math.pow(eAC[2], 2))

    m_BC = math.sqrt(math.pow(eBC[0], 2) + \
                     math.pow(eBC[1], 2) + \
                     math.pow(eBC[2], 2))

    eAB_norm = (eAB[0]/m_AB, eAB[1]/m_AB, eAB[2]/m_AB)
    eAC_norm = (eAC[0]/m_AC, eAC[1]/m_AC, eAC[2]/m_AC)
    eBC_norm = (eBC[0]/m_BC, eBC[1]/m_BC, eBC[2]/m_BC)
    e_vertical = np.cross(eAB_norm, eAC_norm)
    eBO = np.cross(e_vertical, eAB_norm)
    m_BO = math.sqrt(math.pow(eBO[0], 2) + \
                     math.pow(eBO[1], 2) + \
                     math.pow(eBO[2], 2))
    eBO_norm = (eBO[0]/m_BO, eBO[1]/m_BO, eBO[2]/m_BO)
    p_o = None
    p_O1 = (p_B[0]+200*eBO_norm[0], p_B[1]+200*eBO_norm[1], p_B[2]+200*eBO_norm[2])
    p_O2 = (p_B[0]-200*eBO_norm[0], p_B[1]-200*eBO_norm[1], p_B[2]-200*eBO_norm[2])
    costheta = np.dot(eAB_norm, eBC_norm)
    theta = math.acos(costheta)
    # Using theta to filter solution
    if theta >= math.pi/2:
        p_o = None
    # Compare dist to p_O1 and p_O2 to choose circle
    dist1 = math.sqrt(math.pow((p_O1[0]-p_C[0]), 2) + \
                      math.pow((p_O1[1]-p_C[1]), 2) + \
                      math.pow((p_O1[2]-p_C[2]), 2))
    dist2 = math.sqrt(math.pow((p_O2[0]-p_C[0]), 2) + \
                      math.pow((p_O2[1]-p_C[1]), 2) + \
                      math.pow((p_O2[2]-p_C[2]), 2))
    if dist1 < dist2:
        p_o = p_O1
    elif dist1 == dist2:
        p_o = 1
    else:
        p_o = p_O2
    return p_o, e_vertical


def calc_D(x, p_A, p_B, p_C, p_O, e_vertical):
    return np.array([
        math.pow((x[0]-p_O[0]), 2)+math.pow((x[1]-p_O[1]), 2)+math.pow((x[2]-p_O[2]), 2)-200*200,
        (x[0]-p_O[0])*(x[0]-p_C[0])+(x[1]-p_O[1])*(x[1]-p_C[1])+(x[2]-p_O[2])*(x[2]-p_C[2]),
        e_vertical[0]*x[0]+e_vertical[1]*x[1]+e_vertical[2]*x[2]
    ])


if __name__ == '__main__':
    # p_A = (11392.9607416196, 56973.0182393612, 4097.85801775604)
    df = pd.read_excel('test1.xlsx')
    wb = op.load_workbook('circle.xlsx')
    ws = wb.get_active_sheet()
    # list_index = ['237', '233', '598', '561', '485', '230', '204']
    # df_use = df.loc[[237, 233, 598, 561, 485, 230, 204], :]
    point = [(0.0, 50000.0, 5000.0),
             (9019.94516182013, 50399.8851535608, 2569.94886399598),
             (11378.1726511, 51578.6462114014, 8507.54931763946),
             (16754.7303225694, 57311.9910512877, 9794.60069443671),
             (27810.0363906167, 57543.804355148, 5123.83998827498),
             (39684.0766890104, 56894.4082461272, 6550.4738181871),
             (45591.7751823415, 61669.8984619136, 6431.02660574361),
             (55822.0024672237, 64332.3015297792, 9852.15481923928),
             (66726.2371168726, 64030.6571657957, 8825.61539613089),
             (72922.4909659759, 65748.7433586562, 1514.31491761549),
             (80615.3514453143, 61875.6011821751, 4762.59437630022),
             (91535.1784989141, 56331.3622768799, 806.136581376571),
             (100000.0, 59652.34338, 5022.001164)]
    rows = 1
    cols = 5
    list_index = ['0', '40', '578', '417', '80', '170', '282', '598', '561', '406', '248', '46', '612']
    df_use = df.loc[[0, 40, 578, 417, 80, 170, 282, 598, 561, 406, 248, 46, 612], :]
    with open('yuanxin.txt', 'wb') as f:
        # ws.cell(row=rows, column=cols).value = point[0][0]
        # ws.cell(row=rows, column=cols+1).value = point[0][1]
        # ws.cell(row=rows, column=cols+2).value = point[0][2]
        # ws.cell(row=rows+1, column=cols).value = point[1][0]
        # ws.cell(row=rows+1, column=cols+1).value = point[1][1]
        # ws.cell(row=rows+1, column=cols+2).value = point[1][2]
        rows += 2
        for i in range(0, 11):
            df_last = df_use.loc[int(list_index[i]), :]
            df_cur = df_use.loc[int(list_index[i+1]), :]
            df_next = df_use.loc[int(list_index[i+2]), :]
            p_A = (df_last[1], df_last[2], df_last[3])
            p_B = (df_cur[1], df_cur[2], df_cur[3])
            p_C = (df_next[1], df_next[2], df_next[3])
            p_O, e_vertical = calc_O_coord(p_A, p_B, p_C)
            # print(p_O)
            f.write(str(p_O[0]).encode()+b', '+str(p_O[1]).encode()+b', '+str(p_O[2]).encode()+b'\n')
            p_D = fsolve(calc_D, [p_B[0], p_B[1], p_B[2]], args=(p_A, p_B, p_C, p_O, e_vertical))
            print(p_D)
            f.write('切点: '.encode()+str(p_D[0]).encode()+b', '+str(p_D[1]).encode()+b', '+str(p_D[2]).encode()+b'\n')
            ws.cell(row=rows, column=cols).value = p_D[0]
            ws.cell(row=rows, column=cols+1).value = p_D[1]
            ws.cell(row=rows, column=cols+2).value = p_D[2]
            ws.cell(row=rows, column=cols+4).value = p_O[0]
            ws.cell(row=rows, column=cols+5).value = p_O[1]
            ws.cell(row=rows, column=cols+6).value = p_O[2]
            rows += 1
            # ws.cell(row=rows, column=cols).value = point[i+2][0]
            # ws.cell(row=rows, column=cols+1).value = point[i+2][1]
            # ws.cell(row=rows, column=cols+2).value = point[i+2][2]
            # rows += 1
            m_OD = math.sqrt(math.pow((p_D[0]-p_O[0]), 2) + \
                             math.pow((p_D[1]-p_O[1]), 2) + \
                             math.pow((p_D[2]-p_O[2]), 2))
            print(m_OD)
    wb.save('circle.xlsx')
    wb.close()
